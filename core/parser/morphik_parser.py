import asyncio
import io
import logging
import os
import tempfile
from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from docling.datamodel.base_models import InputFormat
from docling.datamodel.pipeline_options import PdfPipelineOptions
from docling.document_converter import DocumentConverter, PdfFormatOption
from httpx import AsyncClient, Timeout

from core.config import get_settings
from core.models.chunk import Chunk
from core.parser.base_parser import BaseParser
from core.parser.video.parse_video import VideoParser, load_config
from core.parser.xml_chunker import XMLChunker
from core.storage.utils_file_extensions import detect_content_type

# Custom RecursiveCharacterTextSplitter replaces langchain's version


logger = logging.getLogger(__name__)


class BaseChunker(ABC):
    """Base class for text chunking strategies"""

    @abstractmethod
    def split_text(self, text: str) -> List[Chunk]:
        """Split text into chunks"""
        pass


class StandardChunker(BaseChunker):
    """Standard chunking using langchain's RecursiveCharacterTextSplitter"""

    def __init__(self, chunk_size: int, chunk_overlap: int):
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""],
        )

    def split_text(self, text: str) -> List[Chunk]:
        return self.text_splitter.split_text(text)


class RecursiveCharacterTextSplitter:
    def __init__(self, chunk_size: int, chunk_overlap: int, length_function=len, separators=None):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.length_function = length_function
        self.separators = separators or ["\n\n", "\n", ". ", " ", ""]

    def split_text(self, text: str) -> list[str]:
        chunks = self._split_recursive(text, self.separators)
        return [Chunk(content=chunk, metadata={}) for chunk in chunks]

    def _split_recursive(self, text: str, separators: list[str]) -> list[str]:
        if self.length_function(text) <= self.chunk_size:
            return [text] if text else []
        if not separators:
            # No separators left, split at chunk_size boundaries
            return [text[i : i + self.chunk_size] for i in range(0, len(text), self.chunk_size)]
        sep = separators[0]
        if sep:
            splits = text.split(sep)
        else:
            # Last fallback: split every character
            splits = list(text)
        chunks = []
        current = ""
        for part in splits:
            add_part = part + (sep if sep and part != splits[-1] else "")
            if self.length_function(current + add_part) > self.chunk_size:
                if current:
                    chunks.append(current)
                current = add_part
            else:
                current += add_part
        if current:
            chunks.append(current)
        # If any chunk is too large, recurse further
        final_chunks = []
        for chunk in chunks:
            if self.length_function(chunk) > self.chunk_size and len(separators) > 1:
                final_chunks.extend(self._split_recursive(chunk, separators[1:]))
            else:
                final_chunks.append(chunk)
        # Handle overlap
        if self.chunk_overlap > 0 and len(final_chunks) > 1:
            overlapped = []
            for i in range(len(final_chunks)):
                chunk = final_chunks[i]
                if i > 0:
                    prev = final_chunks[i - 1]
                    overlap = prev[-self.chunk_overlap :]
                    chunk = overlap + chunk
                overlapped.append(chunk)
            return overlapped
        return final_chunks


class ContextualChunker(BaseChunker):
    """Contextual chunking using LLMs to add context to each chunk"""

    DOCUMENT_CONTEXT_PROMPT = """
    <document>
    {doc_content}
    </document>
    """

    CHUNK_CONTEXT_PROMPT = """
    Here is the chunk we want to situate within the whole document
    <chunk>
    {chunk_content}
    </chunk>

    Please give a short succinct context to situate this chunk within the overall document for the purposes of improving search retrieval of the chunk.
    Answer only with the succinct context and nothing else.
    """

    def __init__(self, chunk_size: int, chunk_overlap: int, anthropic_api_key: str):
        self.standard_chunker = StandardChunker(chunk_size, chunk_overlap)

        # Get the config for contextual chunking
        config = load_config()
        parser_config = config.get("parser", {})
        self.model_key = parser_config.get("contextual_chunking_model", "claude_sonnet")

        # Get the settings for registered models
        from core.config import get_settings

        self.settings = get_settings()

        # Make sure the model exists in registered_models
        if not hasattr(self.settings, "REGISTERED_MODELS") or self.model_key not in self.settings.REGISTERED_MODELS:
            raise ValueError(f"Model '{self.model_key}' not found in registered_models configuration")

        self.model_config = self.settings.REGISTERED_MODELS[self.model_key]
        logger.info(f"Initialized ContextualChunker with model_key={self.model_key}")

    def _situate_context(self, doc: str, chunk: str) -> str:
        import litellm

        # Extract model name from config
        model_name = self.model_config.get("model_name")

        # Create system and user messages
        system_message = {
            "role": "system",
            "content": "You are an AI assistant that situates a chunk within a document for the purposes of improving search retrieval of the chunk.",
        }

        # Add document context and chunk to user message
        user_message = {
            "role": "user",
            "content": f"{self.DOCUMENT_CONTEXT_PROMPT.format(doc_content=doc)}\n\n{self.CHUNK_CONTEXT_PROMPT.format(chunk_content=chunk)}",
        }

        # Prepare parameters for litellm
        model_params = {
            "model": model_name,
            "messages": [system_message, user_message],
            "max_tokens": 1024,
            "temperature": 0.0,
        }

        # Add all model-specific parameters from the config
        for key, value in self.model_config.items():
            if key != "model_name":
                model_params[key] = value

        # Use litellm for completion
        response = litellm.completion(**model_params)
        return response.choices[0].message.content

    def split_text(self, text: str) -> List[Chunk]:
        base_chunks = self.standard_chunker.split_text(text)
        contextualized_chunks = []

        for chunk in base_chunks:
            context = self._situate_context(text, chunk.content)
            content = f"{context}; {chunk.content}"
            contextualized_chunks.append(Chunk(content=content, metadata=chunk.metadata))

        return contextualized_chunks


class MorphikParser(BaseParser):
    """Unified parser that handles different file types and chunking strategies"""

    # Docling converter is expensive to initialize, so we cache it at class level
    _docling_converter: Optional[DocumentConverter] = None

    def __init__(
        self,
        chunk_size: int = 1000,
        chunk_overlap: int = 200,
        assemblyai_api_key: Optional[str] = None,
        anthropic_api_key: Optional[str] = None,
        frame_sample_rate: int = 1,
        use_contextual_chunking: bool = False,
    ):
        # Initialize basic configuration
        self._assemblyai_api_key = assemblyai_api_key
        self._anthropic_api_key = anthropic_api_key
        self.frame_sample_rate = frame_sample_rate

        # Get settings from config
        self.settings = get_settings()
        # Initialize chunker based on configuration
        if use_contextual_chunking:
            self.chunker = ContextualChunker(chunk_size, chunk_overlap, anthropic_api_key)
        else:
            self.chunker = StandardChunker(chunk_size, chunk_overlap)

        # Initialize logger
        self.logger = logging.getLogger(__name__)

        # Setup for API mode parsing
        self._parse_api_endpoints: Optional[List[str]] = None
        self._parse_api_key: Optional[str] = None
        if getattr(self.settings, "PARSER_MODE", "local") == "api":
            if self.settings.MORPHIK_EMBEDDING_API_DOMAIN:
                self._parse_api_endpoints = [
                    f"{ep.rstrip('/')}/parse" for ep in self.settings.MORPHIK_EMBEDDING_API_DOMAIN
                ]
                self._parse_api_key = self.settings.MORPHIK_EMBEDDING_API_KEY
                self.logger.info(f"Parser API mode enabled with {len(self._parse_api_endpoints)} endpoint(s)")

        # Setup for Modal mode parsing
        self._modal_parser_url = getattr(self.settings, "MODAL_PARSER_URL", None)
        self._modal_parser_password = getattr(self.settings, "MODAL_PARSER_PASSWORD", None)
        if getattr(self.settings, "PARSER_MODE", "local") == "modal":
            self.logger.info(f"Modal parser mode enabled at {self._modal_parser_url}")
        self.logger.info(f"PARSER MODE: {getattr(self.settings, 'PARSER_MODE', 'local')}, URL: {self._modal_parser_url}")

    @classmethod
    def _get_docling_converter(cls) -> DocumentConverter:
        """Get or create the cached Docling converter."""
        if cls._docling_converter is None:
            # Configure pipeline options for better performance
            pipeline_options = PdfPipelineOptions()
            # Use fast OCR settings by default
            pipeline_options.do_ocr = True
            pipeline_options.do_table_structure = True

            cls._docling_converter = DocumentConverter(
                format_options={
                    InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options),
                }
            )
        return cls._docling_converter

    def _is_video_file(self, file: bytes, filename: str) -> bool:
        """Check if the file is a video file."""
        try:
            mime_type = detect_content_type(content=file, filename=filename)
            return mime_type.startswith("video/")
        except Exception as e:
            logging.error(f"Error detecting file type: {str(e)}")
            return filename.lower().endswith(".mp4")

    def _is_xml_file(self, filename: str, content_type: Optional[str] = None) -> bool:
        """Check if the file is an XML file."""
        if filename and filename.lower().endswith(".xml"):
            return True
        if content_type and content_type in ["application/xml", "text/xml"]:
            return True
        return False

    def _is_image_file(self, filename: str) -> bool:
        """Check if file is an image."""
        ext = os.path.splitext(filename)[1].lower()
        return ext in {".png", ".jpg", ".jpeg", ".webp"}

    def _is_pdf_file(self, filename: str) -> bool:
        """Check if file is a PDF."""
        return filename.lower().endswith(".pdf")

    def _is_plain_text_file(self, filename: str) -> bool:
        """Check if the file is a plain text file that should be read directly without partitioning."""
        plain_text_extensions = {".txt", ".md", ".markdown", ".json", ".csv", ".tsv", ".log", ".rst", ".yaml", ".yml"}
        lower_filename = filename.lower()
        return any(lower_filename.endswith(ext) for ext in plain_text_extensions)

    # Extensions that openpyxl can handle directly (much faster than Docling)
    _FAST_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}

    def _is_fast_excel_file(self, filename: str) -> bool:
        """Check if the file can be parsed via fast openpyxl path."""
        ext = os.path.splitext(filename.lower())[1]
        return ext in self._FAST_EXCEL_EXTENSIONS

    @staticmethod
    def _parse_excel_to_markdown(file: bytes) -> str:
        """Parse XLSX/XLSM to markdown tables using openpyxl directly.

        Much faster than Docling (sub-second vs minutes) and produces better
        output for RAG because labels and values stay on the same row.
        """
        wb = openpyxl.load_workbook(io.BytesIO(file), data_only=True, read_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[tuple] = []
            for row in ws.iter_rows(values_only=True):
                if all(cell is None for cell in row):
                    continue
                rows.append(row)

            if not rows:
                continue

            parts.append(f"## {sheet_name}\n")

            # Find max columns actually used (trim trailing None columns)
            max_cols = 0
            for row in rows:
                for i in range(len(row) - 1, -1, -1):
                    if row[i] is not None:
                        max_cols = max(max_cols, i + 1)
                        break

            if max_cols == 0:
                continue

            for row_idx, row in enumerate(rows):
                cells = []
                for col_idx in range(max_cols):
                    val = row[col_idx] if col_idx < len(row) else None
                    cell_str = str(val) if val is not None else ""
                    cell_str = cell_str.replace("|", "\\|")
                    cells.append(cell_str)
                parts.append("| " + " | ".join(cells) + " |")
                if row_idx == 0:
                    parts.append("| " + " | ".join(["---"] * max_cols) + " |")

            parts.append("")

        wb.close()
        return "\n".join(parts)

    async def _parse_video(self, file: bytes) -> Tuple[Dict[str, Any], str]:
        """Parse video file to extract transcript and frame descriptions"""
        if not self._assemblyai_api_key:
            raise ValueError("AssemblyAI API key is required for video parsing")

        # Save video to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".mp4") as temp_file:
            temp_file.write(file)
            video_path = temp_file.name

        try:
            # Load the config to get the frame_sample_rate from morphik.toml
            config = load_config()
            parser_config = config.get("parser", {})
            vision_config = parser_config.get("vision", {})
            frame_sample_rate = vision_config.get("frame_sample_rate", self.frame_sample_rate)

            # Process video
            parser = VideoParser(
                video_path,
                assemblyai_api_key=self._assemblyai_api_key,
                frame_sample_rate=frame_sample_rate,
            )
            results = await parser.process_video()

            # Combine frame descriptions and transcript
            frame_text = "\n".join(results.frame_descriptions.time_to_content.values())
            transcript_text = "\n".join(results.transcript.time_to_content.values())
            combined_text = f"Frame Descriptions:\n{frame_text}\n\nTranscript:\n{transcript_text}"

            metadata = {
                "video_metadata": results.metadata,
                "frame_timestamps": list(results.frame_descriptions.time_to_content.keys()),
                "transcript_timestamps": list(results.transcript.time_to_content.keys()),
            }

            return metadata, combined_text
        finally:
            os.unlink(video_path)

    async def _parse_xml(self, file: bytes, filename: str) -> Tuple[List[Chunk], int]:
        """Parse XML file directly using XMLChunker."""
        self.logger.info(f"Processing '{filename}' with dedicated XML chunker.")

        # Get XML parser configuration
        xml_config = {}
        if self.settings and hasattr(self.settings, "PARSER_XML"):
            xml_config = self.settings.PARSER_XML.model_dump()

        # Use XMLChunker to process the XML
        xml_chunker = XMLChunker(content=file, config=xml_config)
        xml_chunks_data = xml_chunker.chunk()

        # Map to Chunk objects
        chunks = []
        for i, chunk_data in enumerate(xml_chunks_data):
            metadata = {
                "unit": chunk_data.get("unit"),
                "xml_id": chunk_data.get("xml_id"),
                "breadcrumbs": chunk_data.get("breadcrumbs"),
                "source_path": chunk_data.get("source_path"),
                "prev_chunk_xml_id": chunk_data.get("prev"),
                "next_chunk_xml_id": chunk_data.get("next"),
            }
            chunks.append(Chunk(content=chunk_data["text"], metadata=metadata))

        return chunks, len(file)

    async def _parse_document_via_api(self, file: bytes, filename: str) -> str:
        """Parse document via remote API (GPU server)."""
        if not self._parse_api_endpoints or not self._parse_api_key:
            raise RuntimeError("Parser API not configured")

        headers = {"Authorization": f"Bearer {self._parse_api_key}"}
        timeout = Timeout(read=300.0, connect=30.0, write=60.0, pool=30.0)

        last_error: Optional[Exception] = None
        for endpoint in self._parse_api_endpoints:
            try:
                async with AsyncClient(timeout=timeout) as client:
                    files = {"file": (filename, file)}
                    data = {"filename": filename}
                    resp = await client.post(endpoint, files=files, data=data, headers=headers)
                    resp.raise_for_status()
                    result = resp.json()
                    return result.get("text", "")
            except Exception as e:
                self.logger.warning(f"Parse API call to {endpoint} failed: {e}")
                last_error = e
                continue

        raise RuntimeError(f"All parse API endpoints failed. Last error: {last_error}")

    async def _parse_document_via_modal(self, file: bytes, filename: str) -> str:
        """Parse document via Modal API with background polling."""
        if not self._modal_parser_url or not self._modal_parser_password:
            raise RuntimeError("Modal parser not configured (MODAL_PARSER_URL or MODAL_PARSER_PASSWORD missing)")

        auth = ("morphik", self._modal_parser_password)
        # Shorter timeout for individual status requests, but we poll
        timeout = Timeout(read=30.0, connect=10.0, write=60.0, pool=30.0)
        base_url = self._modal_parser_url.rstrip("/")

        try:
            async with AsyncClient(timeout=timeout) as client:
                # 1. Start the ingestion task
                files = {"file": (filename, file)}
                data = {"title": filename}
                resp = await client.post(
                    f"{base_url}/ingest",
                    files=files,
                    data=data,
                    auth=auth,
                )
                resp.raise_for_status()
                task_id = resp.json().get("task_id")
                if not task_id:
                    raise RuntimeError("Failed to get task_id from Modal ingestion")

                # 2. Poll for status
                max_retries = 120  # 20 minutes total with 10s wait
                result = None
                for _ in range(max_retries):
                    status_resp = await client.get(f"{base_url}/status/{task_id}", auth=auth)
                    status_resp.raise_for_status()
                    status_data = status_resp.json()

                    if status_data["status"] == "completed":
                        result = status_data["result"]
                        break
                    elif status_data["status"] == "failed":
                        raise RuntimeError(f"Modal parsing task failed: {status_data.get('error')}")

                    await asyncio.sleep(10)

                if not result:
                    raise RuntimeError(f"Modal parsing timed out for {filename} (task_id: {task_id})")

                # Combine pages into markdown
                full_text = []

                # Use summary if available (especially for single images)
                summary = result.get("summary")
                if summary:
                    full_text.append(f"# Summary\n\n{summary}")

                for page in result.get("pages", []):
                    page_num = page.get("page_number")
                    page_content = []

                    page_text = page.get("text", "").strip()
                    if page_text:
                        page_content.append(page_text)

                    for img in page.get("images", []):
                        desc = img.get("description", "").strip()
                        # Avoid duplicating if it matches summary
                        if desc and desc != summary:
                            page_content.append(f"\n> [Image description: {desc}]\n")

                    if page_content:
                        full_text.append(f"## Page {page_num}\n\n" + "\n".join(page_content))

                return "\n\n---\n\n".join(full_text)
        except Exception as e:
            self.logger.error(f"Modal parsing failed for {filename}: {e}")
            raise

    async def _parse_document_local(self, file: bytes, filename: str) -> str:
        """Parse document using local Docling."""
        suffix = os.path.splitext(filename)[1] or ".pdf"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
            temp_file.write(file)
            temp_path = temp_file.name

        try:
            converter = self._get_docling_converter()
            result = converter.convert(temp_path)
            text = result.document.export_to_markdown()

            if not text.strip():
                self.logger.warning(f"Docling returned no text for {filename}")

            return text
        except Exception as e:
            self.logger.error(f"Docling parsing failed for {filename}: {e}")
            return ""
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

    async def _parse_document(self, file: bytes, filename: str) -> Tuple[Dict[str, Any], str]:
        """Parse document using Docling, Modal, or API, or read directly for plain text files."""
        # For plain text files, read directly without parsing
        if self._is_plain_text_file(filename):
            try:
                text = file.decode("utf-8")
            except UnicodeDecodeError:
                text = file.decode("latin-1")
            return {}, text

        # Fast path for Excel files — openpyxl is orders of magnitude faster than Docling
        if self._is_fast_excel_file(filename):
            try:
                text = self._parse_excel_to_markdown(file)
                if text.strip():
                    return {}, text
                self.logger.warning(f"Fast Excel parser returned empty for {filename}, falling back to Docling")
            except Exception as e:
                self.logger.warning(f"Fast Excel parser failed for {filename}: {e}, falling back to Docling")

        # For complex formats, use Modal if configured, otherwise API or local Docling
        parser_mode = getattr(self.settings, "PARSER_MODE", "local")

        # Modal handles PDFs and images
        if parser_mode == "modal" or ((self._is_image_file(filename) or self._is_pdf_file(filename)) and self._modal_parser_url):
            try:
                text = await self._parse_document_via_modal(file, filename)
                return {}, text
            except Exception as e:
                # Fallback to local Docling for PDFs if not in explicit modal mode
                if self._is_pdf_file(filename) and parser_mode != "modal":
                    self.logger.warning(f"Modal parsing failed for {filename}, falling back to local: {e}")
                    text = await self._parse_document_local(file, filename)
                    return {}, text
                else:
                    self.logger.error(f"Modal parsing failed for {filename}: {e}")
                    raise

        if self._parse_api_endpoints:
            try:
                text = await self._parse_document_via_api(file, filename)
                return {}, text
            except Exception as e:
                self.logger.warning(f"API parsing failed, falling back to local: {e}")
                text = await self._parse_document_local(file, filename)
                return {}, text
        else:
            text = await self._parse_document_local(file, filename)
            return {}, text

    async def parse_file_to_text(self, file: bytes, filename: str) -> Tuple[Dict[str, Any], str]:
        """Parse file content into text based on file type"""
        if self._is_video_file(file, filename):
            return await self._parse_video(file)
        elif self._is_xml_file(filename):
            # For XML files, we'll handle parsing and chunking together
            # This method should not be called for XML files in normal flow
            # Return empty to indicate XML files should use parse_and_chunk_xml
            return {}, ""
        return await self._parse_document(file, filename)

    async def parse_and_chunk_xml(self, file: bytes, filename: str) -> List[Chunk]:
        """Parse and chunk XML files in one step."""
        chunks, _ = await self._parse_xml(file, filename)
        return chunks

    def is_xml_file(self, filename: str, content_type: Optional[str] = None) -> bool:
        """Public method to check if file is XML."""
        return self._is_xml_file(filename, content_type)

    async def split_text(self, text: str) -> List[Chunk]:
        """Split text into chunks using configured chunking strategy"""
        return self.chunker.split_text(text)
